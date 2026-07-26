import io
from datetime import datetime
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from app.db.models import BoxType, Brand
from app.core.ledger import get_current_stock


def get_stock_rows(db: Session) -> list[dict]:
    boxes = db.query(BoxType).all()
    rows = []
    for box in boxes:
        stock = get_current_stock(db, box.id)
        brand = db.query(Brand).filter_by(id=box.brand_id).first()
        rows.append({
            "size": box.size_label,
            "brand": brand.name if brand else "—",
            "stock": stock,
            "threshold": box.reorder_threshold,
            "low_stock": stock < box.reorder_threshold,
            "is_custom": box.is_custom,
        })
    rows.sort(key=lambda r: (r["brand"], r["size"]))
    return rows


def build_excel_report(db: Session) -> io.BytesIO:
    rows = get_stock_rows(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    low_stock_fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
    body_font = Font(name="Arial")

    ws["A1"] = f"PackIn.ai — Carton Box Inventory ({datetime.now().strftime('%d-%b-%Y %H:%M')})"
    ws["A1"].font = Font(name="Arial", bold=True, size=13)
    ws.merge_cells("A1:E1")

    headers = ["Box size", "Brand", "Current stock", "Reorder threshold", "Status"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    for i, row in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=row["size"]).font = body_font
        ws.cell(row=i, column=2, value=row["brand"]).font = body_font
        ws.cell(row=i, column=3, value=row["stock"]).font = body_font
        ws.cell(row=i, column=4, value=row["threshold"]).font = body_font
        status = "LOW STOCK" if row["low_stock"] else "OK"
        ws.cell(row=i, column=5, value=status).font = body_font
        if row["low_stock"]:
            for col in range(1, 6):
                ws.cell(row=i, column=col).fill = low_stock_fill

    for col, width in zip("ABCDE", [14, 20, 16, 18, 12]):
        ws.column_dimensions[col].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_pdf_report(db: Session) -> io.BytesIO:
    rows = get_stock_rows(db)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("PackIn.ai — Carton Box Inventory Report", styles["Title"]),
        Paragraph(datetime.now().strftime("%d %b %Y, %H:%M"), styles["Normal"]),
        Spacer(1, 16),
    ]

    data = [["Box size", "Brand", "Stock", "Threshold", "Status"]]
    for row in rows:
        status = "LOW STOCK" if row["low_stock"] else "OK"
        data.append([row["size"], row["brand"], str(row["stock"]), str(row["threshold"]), status])

    table = Table(data, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (2, 0), (-1, -1), "CENTER"),
    ]
    for i, row in enumerate(rows, start=1):
        if row["low_stock"]:
            style.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FCE4E4")))
    table.setStyle(TableStyle(style))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer
